#!/usr/bin/env python
# -*- coding: utf-8 -*-
#Script for extracts
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import os

def sumaValores():
	wb = load_workbook(filename = 'extractos.xlsx')
	ws = wb['Sheet']
	d1 = ws['D1'].value
	d2 = ws['D2'].value
	d3 = ws['D3'].value
	d4 = ws['D4'].value
	d5 = ws['D5'].value
	d6 = ws['D6'].value
	d7 = ws['D7'].value
	d8 = ws['D8'].value
	d9 = ws['D9'].value
	d10 = ws['D10'].value
	d11 = ws['D11'].value
	d12 = ws['D12'].value
	d13 = ws['D13'].value
	d14 = ws['D14'].value
	d15 = ws['D15'].value
	d16 = ws['D16'].value
	d17 = ws['D17'].value
	d18 = ws['D18'].value
	d19 = ws['D19'].value
	d20 = ws['D20'].value
	suma = (d1+d2+d3+d4) - (d5+d6+d7+d8+d9+d10+d11+d12+d13+d14+d15+d16+d17+d18+d19+d20)
	print (suma)
	
	
	
	
	
def modificarCelda(o):
	n = float(input("Ingrese un valor para sumar o restar: "))
	wb = load_workbook('extractos.xlsx')
	ws = wb.active 
	#c = ws.cell(row = o, column = 4)
	#n = float(c.value) + n
	#c.value = n
	n = ws['d'+ str(o)].value + n
	ws['d'+ str(o)].value = n
	wb.save('extractos.xlsx')

def saldoInicial():
	x = float(input("Ingrese el valor inicial: "))
	wb = load_workbook('extractos.xlsx')
	ws = wb.active
	ws['d1'] = x
	wb.save('extractos.xlsx')
	




def gana(n):
	
	if (n == "1"):
		print ("Elegiste opcion depositos")
		depositos()

	elif (n == '2'):
		print ("Elegiste opcion cheques")
		chequeRechazadoG()
				
	elif (n == '3'):
		print ("Elegiste opcion acreed Prestamo")
		acreedPrest()	
	else:
		print ("Ingresando al menu principal...")
	
def depositos():
	modificarCelda(2)

def chequeRechazadoG():
	modificarCelda(3)
	
def acreedPrest():
	modificarCelda(4)





def pierde(n):
	if (n == "1"):
		print ("Elegiste opcion Emisiones")
		emisiones()

	elif (n == '2'):
		print ("Elegiste opcion Embargo")
		embargo()
				
	elif (n == '3'):
		print ("Elegiste opcion Cheque rechazado")
		chequeRechazadoP()	
	elif (n == '4'):
		print ("Elegiste opcion  Tarjeta")
		tarjeta()
	else:
		print ("Ingresando al menu principal...")
	
def emisiones():
	modificarCelda(5)
	
def embargo():
	modificarCelda(6)

def chequeRechazadoP():
	modificarCelda(7)

def cobroPrestamo():
	modificarCelda(8)

def tarjeta():
	modificarCelda(9)







def gastos(n):
	
	if (n == "1"):
		print ("Elegiste opcion No grabado")
		nG()
		
	elif (n == '2'):
		print ("Elegiste opcion Iva 10,5 o 21")
		a = float(input("Ingrese el valor mas alto con el que desea operar: "))
		b = float(input("Ingrese el valor mas bajo con el que desea operar: "))
		ivaOpcion(a,b)
				
	elif (n == '3'):
		print ("Elegiste opcion Percepcion IVA")
		iva()	
	elif (n == '4'):
		print ("Elegiste opcion Ibb CABA")
		ibbCaba()
	elif (n == '5'):
		print ("Elegiste opcion debito")
		deb()
	elif (n == '6'):
		print ("Elegiste opcion credito")
		cred()
	elif(n == '7'):
		print ("Elegiste opcion sellos")
		sellos()
	elif (n == '8'):
		print("Elegiste opcion Sircreb")
		sircreb()
	elif (n == '9'):
		print("Elegiste opcion Afip Otros")
		afipOtros()
	else:
		print ("Ingresando al menu principal...")

def ivaOpcion(a,b):
	c10 = 0.105
	c21 = 0.21
	if ((a * c10 - b) <= 0.2 and (a * c10) -b >= -0.2):
		print ("Iva 10,5%")
		wb = load_workbook('extractos.xlsx')
		ws = wb.active 
		a = ws['d11'].value + a
		ws['d11'].value = a
		b = ws['d13'].value +b
		ws['d13'].value = b
		wb.save('extractos.xlsx')
		
	elif ((a * c21) - b <= 0.2 and (a * c21) - b >= -0.2):
		print ("Iva 21%")
		wb = load_workbook('extractos.xlsx')
		ws = wb.active 
		a = ws['d12'].value + a
		ws['d12'].value = a
		b = ws['d14'].value +b
		ws['d14'].value = b
		wb.save('extractos.xlsx')
	else:
		print ("Invalido")

def nG():
	modificarCelda(10)

def iva():
	modificarCelda(15)

def ibbCaba():
	modificarCelda(16)

def deb():
	modificarCelda(17)

def cred():
	modificarCelda(18)
	
def sellos():
	modificarCelda(19)

def sircreb():
	modificarCelda(20)

def afipOtros():
	modificarCelda(21)




def exitProgram():
	
	while True:
		m = input("Elegiste salir del programa, estas seguro? y/n: ")
		if m == 'y':
			exit()
		elif m == 'n': 
			print("Elegiste no salir del programa")
			break
		else:
			print ("Ingrese una opcion valida")
	





filename = "extractos.xlsx"


if os.path.isfile(filename):
	print("Cargando documento")
	saldoInicial()

else:
	print("Documento inexistente por favor crea el documento correspondiente")
	exit()


while True:
	sumaValores()
	opc = input ("""
	Elija la opcion que desee:
	Opcion '1' = Recibir
	Opcion '2' = Pagar
	Opcion '3' = Gastos bcos
	Opcion '4'  = Salir
	""")
	
	if (opc == '1'):
		print("Elegiste opc Recibir")
		n = input("""
		Elija la opcion que desee: 
		Opcion '1' = Depositos
		Opcion '2' = Cheque Rechazado
		Opcion '3' = Acreed prestamo
		Cualquier otra opcion = Menu principal
		""")
		gana(n)
		
		
	elif (opc == '2'):
		print ("Elegiste opc Pagar") 
		n= input("""
		Elija la opcion que desee:
		Opcion '1' = Emisiones
		Opcion '2' = Embargo
		Opcion '3' = Cheques
		Opcion '4' = Tarjeta
		Cualquier otra opcion = Menu principal
		""")
		pierde(n)
		
		
	elif (opc == '3'):
		print ("Elegiste opc Gastos bcos")
		n = input("""
		Opcion '1' = No grabado
		Opcion '2' = Iva 10,5 o 21
		Opcion '3' = Percepcion IVA
		Opcion '4' = IBB Caba
		Opcion '5' = Debito 
		Opcion '6' = Credito
		Opcion '7' = Sellos 
		Opcion '8' = Sircreb
		Opcion '9' = Afip Otros
		Cualquier otra opcion = Salir al menu principal
		""")
		gastos(n)
		
		
	elif (opc == '4'):
		exitProgram()
	else:
		print("Opcion invalida, ingrese una correcta")
		
	





